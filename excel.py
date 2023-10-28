from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from PIL import Image


def getColour(cols):
    hexVals = ["0","1","2","3","4","5","6","7","8","9","A","B","C","D","E","F"]
    hexCol = 'FF'

    for col in cols:
        hexCol += hexVals[col//16] + hexVals[col % 16]
        
    return hexCol



def getX(x):
    letter = ''
    while x > 25:   
        letter += chr(65 + int((x)/26) - 1)
        x = x - (int((x)/26))*26
    letter += chr(65 + (int(x)))
    return letter

wb = Workbook()
ws = wb.active



img = Image.open("img1.jpg")
im = img.load()
for x in range(img.size[0]):
    for y in range(img.size[1]):

        hexCol = getColour(im[x,y])        
        excelCol = PatternFill(start_color=hexCol,
            end_color=hexCol,
            fill_type='solid')
        
        cell = getX(x) + str(y+1)
        ws[cell].fill = excelCol

wb.save("test.xlsx")